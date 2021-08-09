from openpyxl import load_workbook
from openpyxl import Workbook
from tkinter import messagebox
import os
import tkinter
import pathlib
import sys
import shutil
import re


if not os.path.exists('settingData.txt'):
    if not os.path.exists('data'):
        os.mkdir('data')

    path = os.getcwd()
    Filefolder = pathlib.Path(path)
    TXTfolder = pathlib.Path(str(Filefolder) + '/data')
    Filepath = [str(p) for p in TXTfolder.glob("*.txt") if p.is_file()]
    FILE = [p.name for p in TXTfolder.iterdir() if p.is_file]
    #開始行数と終了行数を指定するテキストファイルが存在しない場合作成する
    #settingData.txtが存在しな状態で実行することで規定フォーマットのテキストファイルが生成されます

    f = open(str(path) + '/settingData.txt' , 'w' , encoding='utf-8', newline='\n')
    f.write('開始行数指定:22\n終了行数指定:37\nExcellのファイル名:test\nセルの優先順位:B>F>C>G>D>H>E>I>J>K>L>M\n読み込み列:2\nテンプレートExcel名:BER-template'
    + '\n\n-----------------------------------------\n'
    + 'SAMPLE\n開始行数指定:' + '22,44,68,27' + '\n終了行数指定:' + '26,56,76,36' + '\n＊行数は1から数え始めます\nExcellのファイル名:'
    + 'sample' + '\n＊.xlsxを記入する必要はありません\nセルの優先順位:B>F>C>G>D>H>E>I>J>K>L>M\n読み込み列:2\nテンプレートExcel名:BER-template'
    + '\n-----------------------------------------')
    f.close()


else:
    path = os.getcwd()
    Filefolder = pathlib.Path(path)
    TXTfolder = pathlib.Path(str(Filefolder) + '/data')
    Filepath = [str(p) for p in TXTfolder.glob("*.txt") if p.is_file()]
    FILE = [p.name for p in TXTfolder.iterdir() if p.is_file]

    if not os.path.exists('data'):
        os.mkdir('data')
        messagebox.showinfo('確認','dataフォルダを生成しました。フォルダ内にデータファイルを保存してください')
    elif len(FILE) == 0:
        messagebox.showinfo('警告','dataフォルダ内にテキストファイルがありません')
    else:
        #開始行と終了行の読み取り
        with open('settingData.txt' , 'r' , encoding="utf-8") as f:
            SET = f.readlines()
            START = re.split('[:|,]' , SET[0])
            del START[0]
            if len(START) == 1:
                for i in range(len(Filepath)):
                    START.append(START[0])
            END = re.split('[:|,]' , SET[1])
            del END[0]
            if len(END) == 1:
                for i in range(len(Filepath)):
                    END.append(END[0])
            OLCOLUMN = re.split('[:|,]' , SET[4])
            del OLCOLUMN[0]
            if len(OLCOLUMN) == 1:
                for i in range(len(Filepath)):
                    OLCOLUMN.append(OLCOLUMN[0])
            Filename = SET[2].replace("\n","").replace("\t","").replace("Excellのファイル名:","")
            CELL_ALL = SET[3].replace("セルの優先順位:","")
            CELL = CELL_ALL.split('>')
            tempExcel = SET[5].replace("\n","").replace("\t","").replace("テンプレートExcel名:","")
            f.close()

        
        CELL_culumn = ['A','B','C','D','E','F','G','H','I','J','K','M']
        row_2 = []
        #ファイル名が以前と同じ時2行目をコピー
        if os.path.exists('./Excel/' + str(Filename) + '.xlsx'):
            Eb = load_workbook('./Excel/' + str(Filename) + '.xlsx')
            Es = Eb.worksheets[0]
            #B2:M2までの対応です。テンプレートを変える際には以下を変更してください
            for IO in range(11):
                LL = IO + 1
                LO = CELL_culumn[LL] + str(2)
                PO = Es[LO].value
                row_2.append(PO)

        #テンプレートの複製
        shutil.copy(tempExcel + '.xlsx','a.xlsx')
        if not os.path.exists('Excel'):
            os.mkdir('Excel')
        if not os.path.exists('usedData'):
            os.mkdir('usedData')

        #Excellへの記入
        wb = load_workbook(filename = 'a.xlsx')
        ws = wb.worksheets[0]
        if os.path.exists('./Excel/' + str(Filename) + '.xlsx'):
            for q in range(len(row_2)):
                r = q + 2
                ws.cell(row = 2 , column = r).value = row_2[q]

        for i in range(0 , len(Filepath)):  
            epx_data = []  
            #必要データを抽出
            with open(str(path) + '/data/' + FILE[i] , 'r' , encoding="utf-8") as f:
                elementary_experiment_data = f.readlines()
                START_SET = int(START[i]) - 1
                END_SET = int(END[i])
                OLCOLUMN_SET = int(OLCOLUMN[i]) - 1
                experiment_data = elementary_experiment_data[START_SET : END_SET]
                f.close()
                b = 0
            
            for d in range(len(experiment_data)):
                epx = experiment_data[d].split('\t')
                epx_data.append(epx[OLCOLUMN_SET])

            for d in range(len(epx_data)):
                col = CELL[i]
                ro = d + 3
                writing_cell = col + str(ro)
                try:
                    float(epx_data[d])
                except ValueError:
                    ws[writing_cell] = float(0.00000000000000000001)
                else:
                    if float(epx_data[d]) = 0:
                        ws[writing_cell] = float(0.00000000000000000001)
                    else:
                        ws[writing_cell] = float(epx_data[d])
        
            
            #shutil.move(str(path) + '/data/' + FILE[i] , str(path) + '/usedData/')
        wb.save('./Excel/' + str(Filename) + '.xlsx')
        os.remove('a.xlsx')